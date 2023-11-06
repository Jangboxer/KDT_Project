import openpyxl
import calendar
# datalist=[["날짜(월.일)","거리(km)","시간(분.초)","평균심박수"]]
datalist=[]
def menu():
    print("="*30)
    print('\033[47m\033[103m'+"\t러닝 프로그램"+'\033[0m')
    print("1. 활동저장")
    print("2. 활동보기 월별")
    print("3. 활동보기 전체")
    print("4. 데이버분석 심박존 피드백")
    print("5. 심박기반 마라톤 예측 수행시간")
    print("0. 엑셀파일로 저장")
    print("7. 종료")
def save(): #1번할때 저장
    print("활동을 저장하세요")
    day=input("날짜를 입력하세요 (월.일)")
    distance=input("거리를 입력하세요(km단위) : ")
    time=input("시간을 입력하세요(분.초)")
    sim=input("평균 심박수를 입력하세요 : ")
    inputlist=[day,distance,time,sim]
    datalist.append(inputlist) #데이터리스트에 입력받은값을 저장.
    pacemin=float(time)/float(distance) #평균페이스
    pacesec=float(pacemin%1*60)
    print(f"거리 {distance}km를 {int(pacemin)}분 {pacesec:0.1f}초 평균페이스로 달렸습니다\n고생하셨습니다!!")
    datalistsave() #사전에 달렸던 기록 저장함수
def write_excel_template(filename, sheetname, listdata): #데이터리스트에 append하여 저장된 데이터리스트를 listdata로 보냄.

    excel_file = openpyxl.Workbook()  # Excel 파일 생성
    excel_sheet = excel_file.active  # Excel 워크시트 생성
    
    if sheetname != '':
        excel_sheet.title = sheetname
    
    for item in listdata:   #데이터리스트 하나하나 꺼내서 시트에 추가하기.
        excel_sheet.append(item)
    
    excel_file.save(filename + ".xlsx") # 파일 이름과 엑셀형식 저장.
    print("입력해주신 데이터들을 엑셀에 저장하였습니다.")
    excel_file.close()  # 파일 닫기
def month():
    load_wb=openpyxl.load_workbook("러닝일지.xlsx") #저장된 엑셀파일을 읽어온다.
    sheet = load_wb.get_sheet_by_name("러닝") #불러온 엑셀 파일에서 시트값을 얻는다.
    input2=input("원하는 월을 선택하세요(예 5.): ")
    text_cal = calendar.TextCalendar(firstweekday=6) #달력모듈
    year = 2023
    input2s=input2.split(".")
    month = int(input2s[0])
    text_cal.prmonth(year, month, w=5, l=2)
    data = []
    month_total_distance=0
    month_total_time=0
    for row in sheet.iter_rows(values_only=True):
        if input2s[0] in row[0] : #row[0]은 A열을 나타냅니다. 5. 을 입력하면 5월에 해당하는 행들만 뽑는다.
            month6=row[0].split(".") #A열은 날짜 6.9로 된것을 스플릿한다.
            time6=row[2].split(".") #C열은 시간 40.1로 된것을 스플릿한다.
            month_total_distance=month_total_distance+float(row[1]) #총 거리를 거리마다 읽어와서 더한다.
            month_total_time=month_total_time+float(row[2]) #총 시간을 시간마다 읽어와서 더한다.
            mtth=month_total_time//60
            mttm=month_total_time%60
            mtts=int(month_total_time%1*100)+1
            if mtts >=60:
                mtts=mtts-60
            else:
                pass    #시간을 깔끔하게 나타나게 하는 식, 평균 페이스도 기록.
            months=month6[0]+"월"+month6[1]+"일 "+row[1]+"KM "+time6[0]+"분"+time6[-1]+"초"
            inttm=int(month_total_time)
            flomm=month_total_time-inttm
            total_sc=(inttm*60)+(flomm*60)
            average_pace_in_seconds = total_sc/month_total_distance  # 초/km

            average_pace_minutes = int(average_pace_in_seconds // 60)  # 평균페이스 분
            average_pace_seconds = int(average_pace_in_seconds % 60)  # 평균페이스 초
            data.append(months) #정리된 러닝기록을 data리스트에 저장.
        
    print(f"{input2}월 저장된 운동기록 : {data}")
    print(f"총 운동한 날짜 {len(data)}일")
    print(f"총거리 : {month_total_distance:0.2f}km\n총 시간 : {int(mtth)}시간 {int(mttm)}분 {mtts}초")
    print(f"평균페이스 {average_pace_minutes}분 {average_pace_seconds }초")
    load_wb.close()  # 파일 닫기
def year():
    load_wb=openpyxl.load_workbook("러닝일지.xlsx") #저장된 엑셀파일을 읽어온다.
    sheet = load_wb.get_sheet_by_name("러닝") #불러온 엑셀 파일에서 시트값을 얻는다.
    data = []
    month_total_distance=0
    month_total_time=0
    for row in sheet.iter_rows(values_only=True):
        month6=row[0].split(".")
        time6=row[2].split(".")
        month_total_distance=month_total_distance+float(row[1])
        month_total_time=month_total_time+float(row[2])
        mtth=month_total_time//60
        mttm=month_total_time%60
        mtts=int(month_total_time%1*100)+1
        if mtts >=60:
            mtts=mtts-60
        else:
            pass
        months=month6[0]+"월"+month6[1]+"일 "+row[1]+"KM "+time6[0]+"분"+time6[-1]+"초"
        inttm=int(month_total_time)
        flomm=month_total_time-inttm
        total_sc=(inttm*60)+(flomm*60)
        average_pace_in_seconds = total_sc/month_total_distance  # 초/km

        average_pace_minutes = int(average_pace_in_seconds // 60)  # 평균페이스 분
        average_pace_seconds = int(average_pace_in_seconds % 60)  # 평균페이스 초
        data.append(months) #정리된 러닝기록을 data리스트에 저장.
        
    print(f"총 운동한 날짜 {len(data)}일")
    print(f"총거리 : {month_total_distance:0.2f}km\n총 시간 : {int(mtth)}시간 {int(mttm)}분 {mtts}초")
    print(f"평균페이스 {average_pace_minutes}분 {average_pace_seconds }초")
    load_wb.close()  # 파일 닫기
def simzone():
    sim1,sim2,sim3,sim4,sim5=0,0,0,0,0
    load_wb=openpyxl.load_workbook("러닝일지.xlsx") #저장된 엑셀파일을 읽어온다.
    sheet = load_wb.get_sheet_by_name("러닝") #불러온 엑셀 파일에서 시트값을 얻는다.
    for row in sheet.iter_rows(values_only=True):
        if float(row[3]) <=139: #로우3은 심박기록인데 이걸 읽어와서 심박존을 생성한다. 
            sim1+=float(row[1]) #심박1존에 뛴 전체 거리
        elif float(row[3])<=157 and float(row[3])>139: #심박2존
            sim2+=float(row[1])
        elif float(row[3])<=167 and float(row[3])>157: #심박3존
            sim3+=float(row[1])
        elif float(row[3])<=179 and float(row[3])>167: #심박4존
            sim4+=float(row[1])
        elif float(row[3])<=187 and float(row[3])>180: #심박5존
            sim5+=float(row[1])
    print('\033[96m'+"="*25+"심박존 그래프"+"="*25+'\033[0m')
    print(f'심박존1 : {"☆"*(int(sim1)//3)}') 
    print(f'심박존2 : {"☆"*(int(sim2)//3)}') #심박존과 거리에 따른 도식화.
    print(f'심박존3 : {"☆"*(int(sim3)//3)}') #눈에 잘보이는 그래프를 활용하여 전체적인 데이터를 빠르게 습득할 수 있다.
    print(f'심박존4 : {"☆"*(int(sim4)//3)}') #거리에서 몫을 구한이유는 거리마다 표시하면 별이 너무 많기에 3을 나눠서 적당량을 그래프로 표현
    print(f'심박존5 : {"☆"*(int(sim5)//3)}')
    if sim1+sim2<sim3+sim4:
        print("고강도훈련을 많이 하였습니다.")
        print("근지구력을 위해 심박수를 낮추고 조깅페이스로 뛰기를 권합니다.")
        print("")
    else:
        print("저강도훈련을 많이 하였습니다.")
        print("폭발적인 에너지를 내기위한 인터벌이나 전력질주 훈련을 권합니다.")
    load_wb.close() #엑셀파일 열었으면 닫기
def predict():
    load_wb=openpyxl.load_workbook("러닝일지.xlsx") #저장된 엑셀파일을 읽어온다.
    sheet = load_wb.get_sheet_by_name("러닝") #불러온 엑셀 파일에서 시트값을 얻는다.
    total_distance=0
    total_time=0
    for row in sheet.iter_rows(values_only=True): #심박4존이 실제 마라톤 기록과 가장 비슷하기때문에 심박4존을 활용.
        if int(row[3])<=179 and int(row[3])>167:
            total_distance+=float(row[1])
            total_time+=float(row[2])
    # print(total_distance)
    # print(total_time)
    averg=total_time/total_distance
    # print(averg*21)
    print("등록된 데이터를 기반으로 마라톤 예상시간을 도출하겠습니다.")
    print(f"◇ 5km마라톤 예상 시간 : {int(averg*5)}분 {float(averg%1)*60:0.0f}초")
    print(f"◇ 10km마라톤 예상 시간 : {int(averg*10)}분 {float(averg%1)*60:0.0f}초")
    print(f"◇ 하프마라톤(21km) 예상 시간 : {(int(averg*21))//60}시간 {int(averg*21)%60+5}분 {float(averg%1)*60:0.0f}초")
    print(f"◇ 풀코스마라톤(42.195km) 예상 시간 : {int((averg*21)//42.195)+1}시간 {int(averg*42.195)%60+10}분 {float(averg%1)*60:0.0f}초")
def datalistsave():